"""
Importar datos desde un archivo Excel (.xlsx) a la base de datos actual.
NO se modifica ni borra nada existente: solo se AGREGAR filas en las tablas
cuyo nombre coincida con el de la hoja. Los datos se insertan tal cual.
Hojas que no coincidan con ninguna tabla se ignoran.

Normalización del nombre de hoja: minúsculas, espacios → guión bajo.
  Ej: "Registro Personal" -> registro_personal, "Base Lockers" -> base_lockers.

Opcional: --export-csv guarda cada hoja en un CSV (sin importar a la BD).

Uso:
  python scripts/import_excel.py datos_otra_area.xlsx
  python scripts/import_excel.py datos.xlsx --export-csv datos_importar
"""
import argparse
import csv
import re
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(SCRIPTS))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env")

try:
    import openpyxl
except ImportError:
    print("Falta la dependencia openpyxl. Instala con: pip install openpyxl")
    sys.exit(1)

from app import create_app, db
from import_datos import (
    IMPORTERS,
    import_registro_personal,
    import_registro_asignaciones,
    import_dotaciones_disponibles,
    import_personal_presupuestado,
    import_locker_disponibles,
    import_historial_retiros,
    import_base_lockers,
    import_base_dotaciones,
)


def normalizar_nombre_hoja(nombre):
    """Convierte nombre de hoja a clave de tabla: minúsculas, espacios -> _."""
    if not nombre:
        return ""
    s = (nombre or "").strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s


def leer_hoja_excel(ws):
    """Convierte una hoja de openpyxl en lista de filas (cada fila = lista de valores)."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Convertir None y números a string para compatibilidad con import_datos
        fila = []
        for cell in row:
            if cell is None:
                fila.append("")
            elif isinstance(cell, (int, float)):
                fila.append(str(cell))
            else:
                fila.append(str(cell).strip() if cell else "")
        rows.append(fila)
    return rows


def exportar_hoja_a_csv(rows, path_csv):
    """Guarda una lista de filas en un archivo CSV (UTF-8 con BOM para Excel)."""
    path_csv = Path(path_csv)
    path_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(path_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for row in rows:
            w.writerow(row)
    return path_csv


def main():
    parser = argparse.ArgumentParser(description="Agregar datos del Excel a la BD actual (solo tablas que coincidan, sin borrar nada)")
    parser.add_argument("excel_file", help="Archivo Excel .xlsx")
    parser.add_argument("--export-csv", metavar="DIR", default=None, help="Solo convertir: guardar cada hoja como CSV en DIR (no importar a BD)")
    args = parser.parse_args()

    path = Path(args.excel_file)
    if not path.exists():
        print(f"No se encuentra el archivo: {path}")
        sys.exit(1)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        print("El archivo debe ser .xlsx (o .xlsm).")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    tablas_conocidas = set(IMPORTERS.keys())

    if args.export_csv:
        # Solo exportar cada hoja a CSV
        out_dir = Path(args.export_csv)
        for name in wb.sheetnames:
            ws = wb[name]
            rows = leer_hoja_excel(ws)
            if not rows:
                continue
            safe_name = re.sub(r'[<>:"/\\|?*]', "_", name.strip()) or "hoja"
            csv_path = out_dir / f"{safe_name}.csv"
            exportar_hoja_a_csv(rows, csv_path)
            print(f"  Exportado: {csv_path}")
        wb.close()
        print("Listo. Para importar a la BD usa: python scripts/import_datos.py <archivo.csv> -t <tabla>")
        return

    app = create_app()
    importados = 0
    ignorados = 0

    with app.app_context():
        for name in wb.sheetnames:
            clave = normalizar_nombre_hoja(name)
            if clave not in tablas_conocidas:
                ignorados += 1
                continue
            ws = wb[name]
            rows = leer_hoja_excel(ws)
            if len(rows) < 2:
                continue
            importer = IMPORTERS[clave]
            try:
                importer(rows, False, app)  # False = solo agregar, nunca reemplazar
                importados += 1
                print(f"  Importado: hoja '{name}' -> tabla '{clave}'")
            except Exception as e:
                print(f"  Error en hoja '{name}' ({clave}): {e}")

    wb.close()
    print(f"\nResumen: {importados} tabla(s) importada(s), {ignorados} hoja(s) ignorada(s) (nombre no coincide).")
    print("Tablas que se importan si la hoja tiene el mismo nombre:", ", ".join(sorted(IMPORTERS.keys())))


if __name__ == "__main__":
    main()
